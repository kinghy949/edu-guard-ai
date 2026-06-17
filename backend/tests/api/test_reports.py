from __future__ import annotations

import io

from openpyxl import load_workbook

from app.services.stats import refresh_snapshots
from tests.conftest import (
    auth_header,
    make_program_with_buckets,
    make_student,
    make_user,
    make_warning,
)


def _staff(db, *, username="rep_staff"):
    u = make_user(db, role="counselor", username=username, password="GoodPass1")
    db.commit()
    return u


def test_export_warnings_xlsx(client, db):
    staff = _staff(db)
    program, _ = make_program_with_buckets(db)
    s = make_student(db, student_no="RP0001", program=program)
    make_warning(db, student=s, level="severe", summary="测试预警 A")
    db.commit()

    r = client.get("/api/v1/reports/warnings.xlsx", headers=auth_header(staff))
    assert r.status_code == 200
    assert "spreadsheet" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["预警明细"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:3] == ("学号", "姓名", "班级")
    assert any(r[0] == "RP0001" for r in rows[1:])


def test_export_completion_xlsx(client, db):
    staff = _staff(db, username="rep_completion")
    program, _ = make_program_with_buckets(db)
    make_student(db, student_no="RP0002", program=program)
    db.commit()
    refresh_snapshots(db)

    r = client.get("/api/v1/reports/completion.xlsx", headers=auth_header(staff))
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["学业完成度"]
    headers = [c.value for c in ws[1]]
    assert "完成度" in headers
    assert any(row[0] == "RP0002" for row in ws.iter_rows(min_row=2, values_only=True))


def test_export_class_summary_xlsx(client, db):
    staff = _staff(db, username="rep_class")
    program, _ = make_program_with_buckets(db)
    make_student(db, student_no="RP0003", program=program, class_name="软工2401")
    db.commit()
    refresh_snapshots(db)
    r = client.get("/api/v1/reports/class-summary.xlsx", headers=auth_header(staff))
    assert r.status_code == 200


def test_student_cannot_export(client, db):
    stu = make_user(db, role="student", username="rep_stu", password="GoodPass1")
    db.commit()
    r = client.get("/api/v1/reports/warnings.xlsx", headers=auth_header(stu))
    assert r.status_code == 403
