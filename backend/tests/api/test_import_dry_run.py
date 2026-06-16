from __future__ import annotations

import io

from openpyxl import load_workbook
from sqlalchemy import select

from app.models.import_batch import ImportBatch, ImportBatchRow
from app.models.student import Student
from tests.conftest import make_user

HEADER = "student_no,name,enroll_year,college,major,class_name\n"


def _staff(client, db, *, username):
    make_user(db, role="counselor", username=username, password="GoodPass1")
    db.commit()
    r = client.post(
        "/api/v1/auth/login",
        data={"username": username, "password": "GoodPass1"},
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _post(client, headers, csv: str, *, dry_run=False, filename="d.csv"):
    files = {"file": (filename, io.BytesIO(csv.encode("utf-8")), "text/csv")}
    url = "/api/v1/imports/students" + ("?dry_run=true" if dry_run else "")
    return client.post(url, files=files, headers=headers)


def test_dry_run_returns_counts_without_business_writes(client, db):
    headers = _staff(client, db, username="dry1")
    csv = HEADER + "D0001,丁一,2024,计算机学院,软件工程,软工2401\n"
    r = _post(client, headers, csv, dry_run=True, filename="dry.csv")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["dry_run"] is True
    assert body["status"] == "dry_run"
    assert body["would_create"] == 1
    assert body["created"] == 1  # 在 savepoint 内确实"创建"过，再回滚

    # 业务表零变化
    assert db.scalar(select(Student).where(Student.student_no == "D0001")) is None

    # 批次记录保留；行级快照仍登记（参考用）
    batch = db.get(ImportBatch, body["batch_id"])
    assert batch.status == "dry_run"
    rows = list(db.scalars(select(ImportBatchRow).where(ImportBatchRow.batch_id == batch.id)))
    assert len(rows) == 1 and rows[0].op == "create"


def test_confirm_after_dry_run_does_write(client, db):
    headers = _staff(client, db, username="dry2")
    csv = HEADER + "D0002,丁二,2024,计算机学院,软件工程,软工2401\n"
    # 预检
    _post(client, headers, csv, dry_run=True)
    assert db.scalar(select(Student).where(Student.student_no == "D0002")) is None
    # 正式提交
    r = _post(client, headers, csv, dry_run=False, filename="confirm.csv")
    assert r.status_code == 200
    assert r.json()["status"] == "completed"
    db.expire_all()
    assert db.scalar(select(Student).where(Student.student_no == "D0002")) is not None


def test_download_error_report_xlsx(client, db):
    headers = _staff(client, db, username="dry3")
    # 包含一条缺 enroll_year 的错误行
    csv = HEADER + "D0003,丁三,,计算机学院,软件工程,软工2401\n"
    body = _post(client, headers, csv).json()
    bid = body["batch_id"]

    resp = client.get(f"/api/v1/imports/batches/{bid}/errors.xlsx", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("行号", "错误信息")
    # 第 2 行业务错误：行号=2（CSV 表头是第 1 行）
    assert rows[1][0] == 2 and "enroll_year" in (rows[1][1] or "")
