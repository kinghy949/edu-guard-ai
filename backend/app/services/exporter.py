"""轻量 openpyxl 封装：把表格规格序列化为可流式下载的 .xlsx。"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class SheetSpec:
    title: str
    headers: list[str]
    rows: list[list[Any]]
    # 列宽自适应时的最大宽度
    max_col_width: int = 40


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_workbook(sheets: list[SheetSpec]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        ws = wb.create_sheet(title=spec.title[:31] or "Sheet")
        ws.append(spec.headers)
        for r in spec.rows:
            ws.append(r)
        # 表头样式
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
        ws.freeze_panes = "A2"
        # 自适应列宽
        for col_idx, header in enumerate(spec.headers, start=1):
            max_len = len(str(header))
            for r in spec.rows:
                v = r[col_idx - 1] if col_idx - 1 < len(r) else ""
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max(max_len + 2, 10), spec.max_col_width,
            )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
